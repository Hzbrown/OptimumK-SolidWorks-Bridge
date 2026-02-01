import pathlib
import json
from datetime import datetime
from typing import Any
import openpyxl



# --- NEW GENERIC PARSER ---
class OptimumSheetParser:
    """
    Generic parser for OptimumK Excel sheets. Handles any number of sheets and sections.
    Coordinates are stored as arrays. Section/block detection is based on spacing.
    """
    def __init__(self, file_path: str):
        self.file_path = pathlib.Path(file_path)
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)

    def parse(self) -> dict:
        result = {}
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            result[sheet_name] = self._parse_sheet(sheet)
        return result

    def _parse_sheet(self, sheet) -> dict:
        rows = list(sheet.iter_rows(values_only=True))
        blocks = self._find_blocks(rows)
        parsed = {}
        for block in blocks:
            block_name, block_rows = block["name"], block["rows"]
            if len(block_rows) > 0:
                header = [str(h).strip() if h is not None else "" for h in block_rows[0]]
                # Wheels block: special handling
                if block_name.lower().startswith("wheels") and "Point Name" in header:
                    parsed[block_name] = self._parse_wheels_block(block_rows)
                # Any block with a Point Name column (except wheels): treat as points
                elif "Point Name" in header:
                    parsed[block_name] = self._parse_points_block(block_rows)
        return parsed

    def _parse_points_block(self, block_rows):
        # Header row (has "Left"/"Right") and optional subheader row (has X/Y/Z)
        hdr0 = [str(h).strip() if h is not None else "" for h in block_rows[0]]
        hdr1 = [str(h).strip() if h is not None else "" for h in block_rows[1]] if len(block_rows) > 1 else []

        # Identify if second row is an X/Y/Z subheader
        has_xyz = hdr1 and ("X" in hdr1 and "Y" in hdr1 and "Z" in hdr1)

        name_idx = hdr0.index("Point Name") if "Point Name" in hdr0 else 1

        left_anchor = hdr0.index("Left") if "Left" in hdr0 else None
        right_anchor = hdr0.index("Right") if "Right" in hdr0 else None

        def find_xyz_cols(anchor, stop_at):
            """
            Find X/Y/Z columns in hdr1 between anchor..stop_at (exclusive).
            If hdr1 doesn't exist, fall back to anchor+1..anchor+3.
            """
            if anchor is None:
                return None, None, None

            if has_xyz:
                lo = anchor
                hi = stop_at if (stop_at is not None and stop_at > anchor) else len(hdr1)
                # search region for X/Y/Z labels
                def find_label(label):
                    for j in range(lo, hi):
                        if hdr1[j] == label:
                            return j
                    return None

                xj = find_label("X")
                yj = find_label("Y")
                zj = find_label("Z")

                # If labels are missing (some files), fall back to contiguous after first found
                if xj is not None and yj is None and xj + 1 < hi:
                    yj = xj + 1
                if xj is not None and zj is None and xj + 2 < hi:
                    zj = xj + 2

                return xj, yj, zj

            # fallback for 1-row header files
            return anchor + 1, anchor + 2, anchor + 3

        left_x_idx, left_y_idx, left_z_idx = find_xyz_cols(left_anchor, right_anchor)
        right_x_idx, right_y_idx, right_z_idx = find_xyz_cols(right_anchor, None)

        # data starts after the header rows (skip subheader if present)
        data_start = 2 if has_xyz else 1

        def f(v):
            try:
                return float(v)
            except (TypeError, ValueError):
                return None

        points = {}
        for row in block_rows[data_start:]:
            if all(cell is None for cell in row):
                continue

            name = row[name_idx] if name_idx < len(row) else None
            if not (name and isinstance(name, str)):
                continue

            x_l = f(row[left_x_idx]) if left_x_idx is not None and left_x_idx < len(row) else None
            y_l = f(row[left_y_idx]) if left_y_idx is not None and left_y_idx < len(row) else None
            z_l = f(row[left_z_idx]) if left_z_idx is not None and left_z_idx < len(row) else None

            x_r = f(row[right_x_idx]) if right_x_idx is not None and right_x_idx < len(row) else x_l
            y_r = f(row[right_y_idx]) if right_y_idx is not None and right_y_idx < len(row) else y_l
            z_r = f(row[right_z_idx]) if right_z_idx is not None and right_z_idx < len(row) else z_l

            # Keep your "None -> 0" behavior:
            points[f"{name}_L"] = [x_l if x_l is not None else 0.0,
                                  y_l if y_l is not None else 0.0,
                                  z_l if z_l is not None else 0.0]
            points[f"{name}_R"] = [x_r if x_r is not None else 0.0,
                                  y_r if y_r is not None else 0.0,
                                  z_r if z_r is not None else 0.0]

        return points

    def _parse_wheels_block(self, block_rows):
        # Parse wheels: parameter_name -> {"left": value, "right": value}
        header = [str(h).strip() if h is not None else "" for h in block_rows[0]]
        name_idx = header.index("Point Name") if "Point Name" in header else 1
        left_idx = header.index("Left") if "Left" in header else 2
        right_idx = header.index("Right") if "Right" in header else 6  # Assuming Right is at column 7 (0-based 6)

        params = {}
        for row in block_rows[1:]:
            if all(cell is None for cell in row):
                continue
            param_name = row[name_idx] if name_idx < len(row) else None
            left_val = row[left_idx] if left_idx < len(row) else None
            right_val = row[right_idx] if right_idx < len(row) else None
            if param_name and isinstance(param_name, str):
                params[param_name] = {"left": left_val, "right": right_val}
        return params

    def _find_blocks(self, rows):
        # A block starts at a row with a non-empty first cell, and ends at the next such row or at the end
        blocks = []
        current_block = None
        for idx, row in enumerate(rows):
            first = row[0] if len(row) > 0 else None
            if first and (not isinstance(first, (int, float))):
                # Start of a new block
                if current_block:
                    current_block["end"] = idx
                    blocks.append({
                        "name": current_block["name"],
                        "rows": rows[current_block["start"]:idx]
                    })
                current_block = {"name": str(first).strip(), "start": idx}
        # Add last block
        if current_block:
            current_block["end"] = len(rows)
            blocks.append({
                "name": current_block["name"],
                "rows": rows[current_block["start"]:]
            })
        return blocks

    def to_json(self, indent: int = 2) -> str:
        return json.dumps(self.parse(), indent=indent, default=str)

    def save_json_per_sheet(self, results_dir: str = "results"):
        """
        Save each sheet as its own JSON file directly in /results/<sheet_name>.json
        """
        base_dir = pathlib.Path(results_dir)
        base_dir.mkdir(parents=True, exist_ok=True)
        parsed = self.parse()
        for sheet_name, sheet_data in parsed.items():
            if 'setup' in sheet_name.lower():
                continue  # Skip setup sheets
            # Clean filename: replace spaces with underscores
            fname = f"{sheet_name.replace(' ', '_')}.json"
            out_path = base_dir / fname
            with open(out_path, "w", encoding="utf-8") as f:
                json.dump(sheet_data, f, indent=2, default=str)
        print(f"Saved {len([s for s in parsed if 'setup' not in s.lower()])} sheets to {base_dir}")

    def parse_reference_distance(self) -> dict:
        """
        Parse the Reference distance from the Excel sheet and return as a dict.
        Looks for a sheet or block named 'Setup' or 'Vehicle Setup', or a cell with 'Reference distance'.
        """
        # First, prioritize sheets with 'setup' in the name
        setup_sheets = [sheet_name for sheet_name in self.workbook.sheetnames if 'setup' in sheet_name.lower()]
        for sheet_name in setup_sheets:
            sheet = self.workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                for i, cell in enumerate(row):
                    if isinstance(cell, str) and 'reference distance' in cell.lower():
                        # Try to get the value from the next cell in the row
                        if i + 1 < len(row):
                            val = row[i + 1]
                            try:
                                val = float(val)
                            except (TypeError, ValueError):
                                val = None
                            return {"Reference distance": val}
        # If not found in setup sheets, search all sheets
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                for i, cell in enumerate(row):
                    if isinstance(cell, str) and 'reference distance' in cell.lower():
                        # Try to get the value from the next cell in the row
                        if i + 1 < len(row):
                            val = row[i + 1]
                            try:
                                val = float(val)
                            except (TypeError, ValueError):
                                val = None
                            return {"Reference distance": val}
        return {"Reference distance": None}

    def save_reference_distance(self, results_dir: str = "results"):
        """
        Save the parsed reference distance to a JSON file directly in the results directory.
        """
        ref_dist = self.parse_reference_distance()
        base_dir = pathlib.Path(results_dir)
        base_dir.mkdir(parents=True, exist_ok=True)
        out_path = base_dir / "Vehicle_Setup.json"
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(ref_dist, f, indent=2, default=str)
        print(f"Reference distance saved to {out_path}")



if __name__ == "__main__":
    parser = OptimumSheetParser(r"Project Files\Excel Vehichle Sheets\Final EV2024.xlsx")
    parser.save_reference_distance("results")
    parser.save_json_per_sheet("results")